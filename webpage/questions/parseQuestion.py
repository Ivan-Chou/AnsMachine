import openpyxl as xl
import json

# =======================================

# global var.

CHAP_AMOUNT = 8

EXL_ELE_TABLE = {
    "Ans":1, 
    "QOrder":2, 
    "Ques":3, 
    "Sel_A":4, 
    "Sel_B":5, 
    "Sel_C":6, 
    "Sel_D":7, 
    "Expl":8, 
    "BeginLine":9, 
    "QPartAmount":10
}# this table stores the correspond column of each part of a question

# =======================================

# functions

def getExlValue(sht, row, column):
    return sht.cell(row = row, column = column).value

def packQuestion(sht, row) -> dict:
    question = dict()

    question["Ans"] = str(getExlValue(sht, row, EXL_ELE_TABLE["Ans"])).replace("\n", "").replace("\t", "")

    qtype = 0

    if(question["Ans"] == "None"):
        return {}

    elif (question["Ans"] == "○"):
        qtype = 1
        question["Ans"] = "0"

    elif(question["Ans"] == "Ｘ"):
        qtype = 1
        question["Ans"] = "1"

    elif(question["Ans"] == "Ａ"):
        qtype = 2
        question["Ans"] = "0"

    elif(question["Ans"] == "Ｂ"):
        qtype = 2
        question["Ans"] = "1"

    elif(question["Ans"] == "Ｃ"):
        qtype = 2
        question["Ans"] = "2"

    elif(question["Ans"] == "Ｄ"):
        qtype = 2
        question["Ans"] = "3"

    else:
        print(f"<ERR> Unknown type of answer: {question['Ans']}")
        return {}

    question["Q_index"] = str(getExlValue(sht, row, EXL_ELE_TABLE["QOrder"])).replace("\n", "").replace("\t", "")
    
    question["Q_statement"] = str(getExlValue(sht, row, EXL_ELE_TABLE["Ques"])).replace("\n", "").replace("\t", "")
    
    if(qtype == 1):
        question["Q_choices"] = [
            "○",
            "Ｘ"
        ]
    else:
        question["Q_choices"] = [
            str(getExlValue(sht, row, EXL_ELE_TABLE["Sel_A"])).replace("\n", "").replace("\t", ""),
            str(getExlValue(sht, row, EXL_ELE_TABLE["Sel_B"])).replace("\n", "").replace("\t", ""),
            str(getExlValue(sht, row, EXL_ELE_TABLE["Sel_C"])).replace("\n", "").replace("\t", ""),
            str(getExlValue(sht, row, EXL_ELE_TABLE["Sel_D"])).replace("\n", "").replace("\t", ""),
        ]
    
    
    question["Expl"] = str(getExlValue(sht, row, EXL_ELE_TABLE["Expl"])).replace("\n", "").replace("\t", "")

    return question

# =======================================

# main
if __name__ == "__main__":
    Qs = [] # Qs[chap][part][Q_index][attr]

    questionFile = xl.load_workbook(f"./XLSX/ch1.xlsx")
    sht = questionFile.worksheets[0]
    questionFile.close()

    for chNow in range(CHAP_AMOUNT):
        sht = None

        Qs.append([])
        
        questionFile = xl.load_workbook(f"./XLSX/ch{chNow + 1}.xlsx")
        sht = questionFile.worksheets[0]
        questionFile.close()

        for partNow in range(2): # 0 -> O/X; 1 -> choice
            Qs[chNow].append([])

            # the "+ 2" is hard-coded for the alignment of (.xlsx)
            partBegin = int(getExlValue(sht, partNow + 2, EXL_ELE_TABLE["BeginLine"])) + 1 # note that partBegin is the row of "第 n 章" => need to + 1
            QAmount = int(getExlValue(sht, partNow + 2, EXL_ELE_TABLE["QPartAmount"])) 

            for qRowOffset in range(QAmount):
                question = packQuestion(sht, partBegin + qRowOffset)

                if(question != {}):
                    Qs[chNow][partNow].append(question)

    # TODO: add img question
    
    with open("../scripts/Qs.js", mode="w") as file:
        file.write(f"let Qs_obj = JSON.parse(`")
        json.dump(Qs, file, ensure_ascii=False)
        file.write("`)")